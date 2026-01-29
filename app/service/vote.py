from typing import List, Optional
from datetime import datetime
from sqlalchemy import func, text
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.model.vote import Vote, VoteEnum
from app.schema.vote import VoteCreate, VoteRead, VoteStats, VoteUpdate, VoteWithMessage
from app.config.database import global_schema

class VoteService:
    def __init__(self, db: Session):
        self.db = db

    def create_vote(self, vote_data: VoteCreate) -> VoteRead:
        """创建新的投票"""
        try:
            # 检查该消息是否已经被投票过（可选业务逻辑）
            existing_vote_db = self.db.query(Vote).filter(
                Vote.message_id == vote_data.message_id
            ).first()

            if existing_vote_db:
                # 如果存在，更新现有投票
                existing_vote_db.set_vote_type(vote_data.vote_type)
                if vote_data.feedback is not None:
                    existing_vote_db.set_feedback_content(feedback_content=vote_data.feedback)
                else:
                    existing_vote_db.set_feedback_content(feedback_content="")
                self.db.commit()
                self.db.refresh(existing_vote_db)
                return VoteRead.model_validate(existing_vote_db)

            # 创建新投票
            vote = Vote(
                message_id=vote_data.message_id,
                vote_type=vote_data.vote_type.value,
                feedback = vote_data.feedback                
            )

            self.db.add(vote)
            self.db.commit()
            self.db.refresh(vote)

            return VoteRead.model_validate(vote)
        except Exception as e:
            self.db.rollback()
            raise e

    def get_vote_by_id(self, vote_id: int) -> Optional[VoteRead]:
        """根据ID获取投票"""
        vote = self.db.query(Vote).filter(Vote.vote_id == vote_id).first()
        if vote:
            return VoteRead.model_validate(vote)
        return None

    def get_votes_by_message(self, message_id: int) -> List[VoteRead]:
        """获取某个消息的所有投票"""
        votes = self.db.query(Vote).filter(Vote.message_id == message_id).all()
        return [VoteRead.model_validate(vote) for vote in votes]

    def get_all_votes(self, page: int = 1, size: int = 10) -> List[VoteRead]:
        """分页获取所有投票"""
        offset = (page - 1) * size
        votes = self.db.query(Vote).offset(offset).limit(size).all()
        return [VoteRead.model_validate(vote) for vote in votes]

    def get_total_votes_count(self) -> int:
        """获取总投票数"""
        return self.db.query(Vote).count()

    def update_vote(self, vote_id: int, vote_data: VoteUpdate) -> VoteRead:
        """更新投票"""
        try:
            vote = self.db.query(Vote).filter(Vote.vote_id == vote_id).first()
            if not vote:
                raise ValueError(f"Vote with ID {vote_id} not found")
            vote.vote_type = vote_data.vote_type.value
            self.db.commit()
            self.db.refresh(vote)

            return VoteRead.model_validate(vote)
        except Exception as e:
            self.db.rollback()
            raise e

    def delete_vote(self, vote_id: int) -> bool:
        """删除投票"""
        try:
            vote = self.db.query(Vote).filter(Vote.vote_id == vote_id).first()
            if not vote:
                raise ValueError(f"Vote with ID {vote_id} not found")

            self.db.delete(vote)
            self.db.commit()

            return True
        except Exception as e:
            self.db.rollback()
            raise e

    def get_vote_stats_by_message(self, message_id: int) -> Optional[VoteStats]:
        """获取某个消息的投票统计"""
        from sqlalchemy import case

        # 构建统计查询
        stats_query = (
            self.db.query(
                Vote.message_id,
                func.sum(case((Vote.vote_type == VoteEnum.good, 1), else_=0)).label('good_count'),
                func.sum(case((Vote.vote_type == VoteEnum.medium, 1), else_=0)).label('average_count'),
                func.sum(case((Vote.vote_type == VoteEnum.bad, 1), else_=0)).label('poor_count'),
                func.count(Vote.vote_id).label('total_count')
            )
            .filter(Vote.message_id == message_id)
            .group_by(Vote.message_id)
            .first()
        )

        if stats_query:
            return VoteStats(
                message_id=stats_query.message_id,
                good_count=int(stats_query.good_count or 0),
                average_count=int(stats_query.average_count or 0),
                poor_count=int(stats_query.poor_count or 0),
                total_count=int(stats_query.total_count or 0)
            )
        return None

    def get_vote_stats_by_type(self, vote_type: VoteEnum) -> int:
        """获取某种投票类型的总数"""
        return self.db.query(Vote).filter(Vote.vote_type == vote_type).count()

    def get_user_vote_for_message(self, message_id: int, user_id: Optional[str] = None) -> Optional[VoteRead]:
        """获取用户对特定消息的投票（如果需要用户关联的话）"""
        # 注意：当前的 Vote 模型没有 user_id 字段，这里作为扩展接口
        # 如果需要用户关联，需要修改模型
        vote = self.db.query(Vote).filter(Vote.message_id == message_id).first()
        if vote:
            return VoteRead.model_validate(vote)
        return None

    def get_votes_with_messages(
        self,
        page: int = 1,
        size: int = 10,
        vote_type: Optional[VoteEnum] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        search_keyword: Optional[str] = None,
        client_type: Optional[str] = None
    ) -> List[VoteWithMessage]:
        """获取带问题和答案的投票列表（支持按类型和时间过滤）"""
        offset = (page - 1) * size


        base_query = f"""
            select 
                a.id as message_id,
                a.chat_id,
                a.content as question,
                user_latest.content as answer,
                a.created_at as created_at,
                coalesce(c.vote_type,'unknown') vote_type,
                c.vote_id,
                c.feedback feedback,
                c.updated_at,
                CASE 
                    WHEN a.metadata_->>'client' = 'web' THEN '网页'
                    WHEN a.metadata_->>'client' = 'h5' THEN 'H5'
                    WHEN a.metadata_->>'client' = 'miniprogram' THEN '小程序'
                    WHEN a.metadata_->>'client' = 'mp' THEN '公众号'
                    WHEN a.metadata_->>'client' = '公积金' THEN '公积金'
                    WHEN a.metadata_->>'client' = 'rexian' THEN '热线'
                    ELSE a.metadata_->>'client'
                END as client_type
            from {global_schema}.messages a
            left join lateral (
                select id,chat_id,message_role_enum,content,created_at 
                from {global_schema}.messages 
                where chat_id = a.chat_id
                and message_role_enum = 'assistant'
                and id > a.id
                order by created_at asc limit 1
            ) user_latest ON true
            left join {global_schema}.vote c on user_latest.id = c.message_id 
            where a.message_role_enum = 'user'
        """
        # 构建条件参数
        conditions = []
        params = {"limit": size, "offset": offset}

        if vote_type:
            conditions.append("AND c.vote_type = :vote_type")
            params["vote_type"] = vote_type.value

        if start_date:
            conditions.append("AND a.created_at >= :start_date")
            params["start_date"] = start_date

        if end_date:
            conditions.append("AND a.created_at <= :end_date")
            params["end_date"] = end_date

        if search_keyword:
            conditions.append("AND (a.content ILIKE :search_keyword OR user_latest.content ILIKE :search_keyword)")
            params["search_keyword"] = f"%{search_keyword}%"

        if client_type:
            conditions.append("AND a.metadata_->>'client' = :client_type")
            params["client_type"] = client_type

        # 组装完整查询
        full_query = base_query + " ".join(conditions) + " ORDER BY a.created_at DESC LIMIT :limit OFFSET :offset"

        result = self.db.execute(text(full_query), params)
        rows = result.fetchall()

        return [VoteWithMessage(
            vote_id=row.vote_id,
            message_id=row.message_id,
            vote_type=row.vote_type,
            feedback=row.feedback,
            created_at=row.created_at,
            question=row.question,
            answer=row.answer,
            chat_id=row.chat_id,
            client_type=row.client_type
        ) for row in rows]

    def get_votes_with_messages_count(
        self,
        vote_type: Optional[VoteEnum] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        search_keyword: Optional[str] = None,
        client_type: Optional[str] = None
    ) -> int:
        """获取带问题和答案的投票总数（用于分页）"""

        # 构建基础查询（与 get_votes_with_messages 保持一致）
        base_query = f"""
            SELECT COUNT(DISTINCT a.id) as total
            FROM {global_schema}.messages a
            LEFT JOIN LATERAL (
                SELECT id,chat_id,message_role_enum,content,created_at 
                FROM {global_schema}.messages 
                WHERE chat_id = a.chat_id
                AND message_role_enum = 'assistant'
                and id > a.id
                order by created_at asc limit 1
            ) user_latest ON true
            LEFT JOIN {global_schema}.vote c ON user_latest.id = c.message_id 
            WHERE a.message_role_enum = 'user'
        """

        # 构建条件参数
        conditions = []
        params = {}

        if vote_type:
            conditions.append("AND c.vote_type = :vote_type")
            params["vote_type"] = vote_type.value

        if start_date:
            conditions.append("AND a.created_at >= :start_date")
            params["start_date"] = start_date

        if end_date:
            conditions.append("AND a.created_at <= :end_date")
            params["end_date"] = end_date

        if search_keyword:
            conditions.append("AND (a.content ILIKE :search_keyword OR user_latest.content ILIKE :search_keyword)")
            params["search_keyword"] = f"%{search_keyword}%"

        if client_type:
            conditions.append("AND a.metadata_->>'client' = :client_type")
            params["client_type"] = client_type

        # 组装完整查询
        full_query = base_query + " ".join(conditions)

        result = self.db.execute(text(full_query), params)
        row = result.fetchone()

        return int(row.total) if row else 0

    def get_votes_with_messages_by_chat(self, chat_id: str) -> List[VoteWithMessage]:
        """根据聊天ID获取带问题和答案的投票列表"""
        query = text(f"""
            SELECT
                a.vote_id,
                a.message_id,
                a.vote_type,
                a.updated_at,
                user_latest.content as question,
                b.content as answer,
                b.chat_id,
                CASE 
                    WHEN a.metadata_->>'client' = 'web' THEN '网页'
                    WHEN a.metadata_->>'client' = 'h5' THEN 'H5'
                    WHEN a.metadata_->>'client' = 'miniprogram' THEN '小程序'
                    WHEN a.metadata_->>'client' = 'mp' THEN '公众号'
                    WHEN a.metadata_->>'client' = '公积金' THEN '公积金'
                    WHEN a.metadata_->>'client' = 'rexian' THEN '热线'
                    ELSE a.metadata_->>'client'
                END as client_type
            FROM {global_schema}.vote a
            LEFT JOIN {global_schema}.messages b ON a.message_id = b.id
            LEFT JOIN LATERAL (
                SELECT *
                FROM {global_schema}.messages
                WHERE chat_id = b.chat_id
                AND message_role_enum = 'user'
                AND id < b.id     
                ORDER BY created_at DESC
                LIMIT 1
            ) user_latest ON true
            WHERE b.chat_id = :chat_id
            ORDER BY a.updated_at DESC
        """)

        result = self.db.execute(query, {"chat_id": chat_id})
        rows = result.fetchall()

        return [VoteWithMessage(
            vote_id=row.vote_id,
            message_id=row.message_id,
            vote_type=row.vote_type,
            feedback=row.feedback,
            client_type=row.client_type,
            created_at=row.created_at,
            question=row.question,
            answer=row.answer,
            chat_id=row.chat_id
        ) for row in rows]

    def get_all_votes_with_messages(
        self,
        vote_type: Optional[VoteEnum] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        search_keyword: Optional[str] = None,
        client_type: Optional[str] = None
    ) -> List[VoteWithMessage]:
        """获取所有带问题和答案的投票列表（不分页，用于导出）"""
        # 构建基础查询（与 get_votes_with_messages 保持一致，但去掉分页）
        base_query = f"""
            select 
                a.id as message_id,
                a.chat_id,
                a.content as question,
                user_latest.content as answer,
                a.created_at as created_at,
                coalesce(c.vote_type,'unknown') vote_type,
                c.vote_id,
                c.feedback feedback,
                c.updated_at,
                CASE 
                    WHEN a.metadata_->>'client' = 'web' THEN '网页'
                    WHEN a.metadata_->>'client' = 'h5' THEN 'H5'
                    WHEN a.metadata_->>'client' = 'miniprogram' THEN '小程序'
                    WHEN a.metadata_->>'client' = 'mp' THEN '公众号'
                    WHEN a.metadata_->>'client' = '公积金' THEN '公积金'
                    WHEN a.metadata_->>'client' = 'rexian' THEN '热线'
                    ELSE a.metadata_->>'client'
                END as client_type
            from {global_schema}.messages a
            left join lateral (
                select id,chat_id,message_role_enum,content,created_at 
                from {global_schema}.messages 
                where chat_id = a.chat_id
                and message_role_enum = 'assistant'
                and id > a.id
                order by created_at asc limit 1
            ) user_latest ON true
            left join {global_schema}.vote c on user_latest.id = c.message_id 
            where a.message_role_enum = 'user'
        """
        # 构建条件参数
        conditions = []
        params = {}

        if vote_type:
            conditions.append("AND c.vote_type = :vote_type")
            params["vote_type"] = vote_type.value

        if start_date:
            conditions.append("AND a.created_at >= :start_date")
            params["start_date"] = start_date

        if end_date:
            conditions.append("AND a.created_at <= :end_date")
            params["end_date"] = end_date

        if search_keyword:
            conditions.append("AND (a.content ILIKE :search_keyword OR user_latest.content ILIKE :search_keyword)")
            params["search_keyword"] = f"%{search_keyword}%"

        if client_type:
            conditions.append("AND a.metadata_->>'client' = :client_type")
            params["client_type"] = client_type

        # 组装完整查询（不分页）
        full_query = base_query + " ".join(conditions) + " ORDER BY a.created_at DESC"

        result = self.db.execute(text(full_query), params)
        rows = result.fetchall()

        return [VoteWithMessage(
            vote_id=row.vote_id,
            message_id=row.message_id,
            vote_type=row.vote_type,
            feedback=row.feedback,
            created_at=row.created_at,
            question=row.question,
            answer=row.answer,
            chat_id=row.chat_id,
            client_type=row.client_type
        ) for row in rows]

    def export_votes_to_excel(
        self,
        vote_type: Optional[VoteEnum] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        search_keyword: Optional[str] = None,
        client_type: Optional[str] = None
    ) -> BytesIO:
        """导出投票数据到Excel"""
        # 获取所有数据
        votes = self.get_all_votes_with_messages(
            vote_type=vote_type,
            start_date=start_date,
            end_date=end_date,
            search_keyword=search_keyword,
            client_type=client_type
        )

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "投票数据"

        # 定义表头（添加请求来源列）
        headers = [
            "投票类型",
            "消息ID",
            "用户问题",
            "AI回答",
            "反馈内容",
            "请求来源",
            "消息时间"
        ]

        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 写入数据
        vote_type_map = {
            "good": "好评",
            "medium": "中评",
            "bad": "差评",
            "unknown": "未知"
        }

        for row_idx, vote in enumerate(votes, start=2):
            # 确保所有值都转换为Excel可以处理的类型（字符串、数字、日期）
            # 处理投票类型：如果是枚举对象，获取其value；如果是字符串，直接使用
            vote_type_value = vote.vote_type.value if hasattr(vote.vote_type, 'value') else str(vote.vote_type) if vote.vote_type else "unknown"
            ws.cell(row=row_idx, column=1, value=vote_type_map.get(vote_type_value, "未知"))
            ws.cell(row=row_idx, column=2, value=int(vote.message_id) if vote.message_id else "")
            ws.cell(row=row_idx, column=3, value=str(vote.question) if vote.question else "")
            ws.cell(row=row_idx, column=4, value=str(vote.answer) if vote.answer else "")
            ws.cell(row=row_idx, column=5, value=str(vote.feedback) if vote.feedback else "")
            ws.cell(row=row_idx, column=6, value=str(vote.client_type) if vote.client_type else "")
            # 格式化日期时间（第7列）
            if vote.created_at:
                if isinstance(vote.created_at, str):
                    dt = datetime.fromisoformat(vote.created_at.replace('Z', '+00:00'))
                else:
                    dt = vote.created_at
                ws.cell(row=row_idx, column=7, value=dt.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                ws.cell(row=row_idx, column=7, value="")

        # 设置列宽（添加请求来源列，调整为7列）
        column_widths = [12, 12, 50, 80, 50, 15, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 设置文本换行和对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
    
    