from sqlalchemy.orm import Session
from typing import List, Optional
import os
import uuid
from fastapi import UploadFile, HTTPException
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.model.feedback import Feedback
from app.schema.feedback import FeedbackCreate, FeedbackUpdate, ImageInfo, ImageUploadResponse


class FeedbackService:
    def __init__(self, db: Session):
        self.db = db

    def create_feedback(self, feedback_data: FeedbackCreate) -> Feedback:
        """创建反馈"""
        print(f"服务层收到的数据: {feedback_data}")  # 调试日志

        try:
            feedback = Feedback(
                content=feedback_data.content,
                phone=feedback_data.phone,
                images=[img.dict() for img in feedback_data.images] if feedback_data.images else None
            )

            print(f"创建的反馈对象: {feedback}")  # 调试日志

            self.db.add(feedback)
            self.db.commit()
            self.db.refresh(feedback)

            print(f"反馈创建成功: {feedback.id}")  # 调试日志
            return feedback

        except Exception as e:
            print(f"数据库操作失败: {e}")  # 调试日志
            import traceback
            traceback.print_exc()  # 打印详细错误信息
            self.db.rollback()
            raise e

    def get_feedback_by_id(self, feedback_id: int) -> Optional[Feedback]:
        """根据ID获取反馈"""
        return self.db.query(Feedback).filter(Feedback.id == feedback_id).first()

    def _apply_filters(
        self,
        query,
        content_keyword: Optional[str] = None,
        phone_keyword: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ):
        """在查询上应用通用过滤条件"""
        if content_keyword:
            query = query.filter(Feedback.content.ilike(f"%{content_keyword}%"))
        if phone_keyword:
            query = query.filter(Feedback.phone.ilike(f"%{phone_keyword}%"))
        if start_date:
            query = query.filter(Feedback.created_time >= start_date)
        if end_date:
            query = query.filter(Feedback.created_time <= end_date)
        return query

    def get_all_feedbacks(
        self,
        page: int = 1,
        size: int = 10,
        content_keyword: Optional[str] = None,
        phone_keyword: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> List[Feedback]:
        """分页获取反馈并支持条件筛选"""
        offset = (page - 1) * size
        query = self._apply_filters(
            self.db.query(Feedback),
            content_keyword=content_keyword,
            phone_keyword=phone_keyword,
            start_date=start_date,
            end_date=end_date,
        )
        return (
            query.order_by(Feedback.created_time.desc())
            .offset(offset)
            .limit(size)
            .all()
        )

    def get_total_feedbacks_count(
        self,
        content_keyword: Optional[str] = None,
        phone_keyword: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> int:
        """获取满足条件的反馈总数"""
        query = self._apply_filters(
            self.db.query(Feedback),
            content_keyword=content_keyword,
            phone_keyword=phone_keyword,
            start_date=start_date,
            end_date=end_date,
        )
        return query.count()

    def update_feedback(self, feedback_id: int, feedback_data: FeedbackUpdate) -> Optional[Feedback]:
        """更新反馈"""
        feedback = self.get_feedback_by_id(feedback_id)
        if not feedback:
            return None

        update_data = feedback_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            if field == 'images' and value is not None:
                setattr(feedback, field, [img.dict() for img in value])
            else:
                setattr(feedback, field, value)

        self.db.commit()
        self.db.refresh(feedback)
        return feedback

    def delete_feedback(self, feedback_id: int) -> bool:
        """删除反馈"""
        feedback = self.get_feedback_by_id(feedback_id)
        if not feedback:
            return False

        self.db.delete(feedback)
        self.db.commit()
        return True

    def export_feedbacks_to_excel(
        self,
        content_keyword: Optional[str] = None,
        phone_keyword: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
    ) -> BytesIO:
        """根据条件导出反馈数据为Excel"""
        query = self._apply_filters(
            self.db.query(Feedback),
            content_keyword=content_keyword,
            phone_keyword=phone_keyword,
            start_date=start_date,
            end_date=end_date,
        ).order_by(Feedback.created_time.desc())

        feedbacks = query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "反馈数据"

        headers = ["ID", "反馈内容", "手机号", "状态", "创建时间", "更新时间"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx, fb in enumerate(feedbacks, start=2):
            ws.cell(row=row_idx, column=1, value=fb.id)
            ws.cell(row=row_idx, column=2, value=str(fb.content or ""))
            ws.cell(row=row_idx, column=3, value=str(fb.phone or ""))
            ws.cell(row=row_idx, column=4, value=str(fb.status or ""))
            ws.cell(
                row=row_idx,
                column=5,
                value=fb.created_time.strftime("%Y-%m-%d %H:%M:%S") if fb.created_time else "",
            )
            ws.cell(
                row=row_idx,
                column=6,
                value=fb.updated_time.strftime("%Y-%m-%d %H:%M:%S") if fb.updated_time else "",
            )

        column_widths = [10, 60, 18, 10, 22, 22]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def upload_image(self, file: UploadFile, upload_dir: str = "uploads/feedback") -> ImageUploadResponse:
        """上传图片文件"""
        # 验证文件类型
        allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/gif", "image/webp"]
        if file.content_type not in allowed_types:
            raise HTTPException(
                status_code=400,
                detail=f"不支持的文件类型: {file.content_type}。支持的类型: {', '.join(allowed_types)}"
            )

        # 验证文件大小 (最大5MB)
        max_size = 5 * 1024 * 1024  # 5MB
        file_content = await file.read()
        if len(file_content) > max_size:
            raise HTTPException(
                status_code=400,
                detail=f"文件大小超过限制。最大允许: {max_size // (1024 * 1024)}MB"
            )

        # 创建上传目录
        os.makedirs(upload_dir, exist_ok=True)

        # 生成唯一文件名
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else ''
        unique_filename = f"{uuid.uuid4()}.{file_extension}"
        file_path = os.path.join(upload_dir, unique_filename)

        # 保存文件
        with open(file_path, "wb") as buffer:
            buffer.write(file_content)

        # 构建文件访问URL (相对路径)
        file_url = f"/{upload_dir}/{unique_filename}"

        return ImageUploadResponse(
            url=file_url,
            filename=file.filename,
            size=len(file_content),
            content_type=file.content_type,
            path=file_path
        )